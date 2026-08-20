"""千牛模板列定义：与 5.15新表格.xlsx 第一行表头对齐。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .parser import ParseResult

# 与 5.15新表格.xlsx 默认表头一致（模板读取失败时兜底）
DEFAULT_TEMPLATE_HEADERS = [
    "收件人(必填)",
    "手机号(必填)",
    "收货地址(必填)",
    "平台订单号（非必填）",
    "商品信息(非必填)",
    "规格信息(非必填)",
    "商品数量(非必填)",
    "重量kg(非必填)",
    "备注(非必填)",
]

COLUMN_COUNT = 9
COLS = "ABCDEFGHI"


def read_template_headers(template_path: str | Path) -> list[str]:
    """从模板 xlsx 第一行读取 A~I 表头，保证与千牛模板完全一致。"""
    path = Path(template_path)
    if not path.exists():
        return DEFAULT_TEMPLATE_HEADERS.copy()
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers: list[str] = []
        for col in range(1, COLUMN_COUNT + 1):
            val = ws.cell(1, col).value
            headers.append(str(val).strip() if val is not None else DEFAULT_TEMPLATE_HEADERS[col - 1])
        wb.close()
        if len(headers) == COLUMN_COUNT and all(headers):
            return headers
    except Exception:
        pass
    return DEFAULT_TEMPLATE_HEADERS.copy()


def remark_display(item: ParseResult) -> str:
    """I 列展示：用户备注优先，否则解析错误提示。"""
    if item.remark:
        return item.remark
    return item.error or ""


def result_to_row(item: ParseResult, extract_order_no_to_d: bool = False) -> list[str]:
    """ParseResult → 预览/TSV 九列数据。"""
    return [
        item.name,
        item.phone,
        item.address,
        item.order_no if extract_order_no_to_d else "",
        item.product_info,
        item.spec_info,
        item.quantity,
        item.weight,
        remark_display(item),
    ]


def sync_row_to_result(item: ParseResult, row: list, extract_order_no_to_d: bool = False) -> None:
    """预览表编辑 → 写回 ParseResult。"""
    cells = (list(row) + [""] * COLUMN_COUNT)[:COLUMN_COUNT]
    item.name = str(cells[0] or "").strip()
    item.phone = str(cells[1] or "").strip()
    item.address = str(cells[2] or "").strip()
    if extract_order_no_to_d:
        item.order_no = str(cells[3] or "").strip()
    item.product_info = str(cells[4] or "").strip()
    item.spec_info = str(cells[5] or "").strip()
    item.quantity = str(cells[6] or "").strip()
    item.weight = str(cells[7] or "").strip()
    item.remark = str(cells[8] or "").strip()
