"""自动化测试：读取测试样例并验证解析+导出。"""

from pathlib import Path

from openpyxl import load_workbook

from src.batch_io import read_text_file
from src.exporter import export_to_xlsx
from src.keywords import apply_keyword_hits
from src.parser import parse_batch, split_records
from src.validators import apply_phone_validation

ROOT = Path(__file__).resolve().parent
SAMPLES_FILE = ROOT / "测试样例-批量.txt"


def main():
    raw = SAMPLES_FILE.read_text(encoding="utf-8").strip()
    records = split_records(raw)
    print(f"分隔得到 {len(records)} 条记录\n")

    results = parse_batch(raw, use_ai=False)
    apply_phone_validation(results)
    apply_keyword_hits(results, ["驿站", "工商局", "质检局"])

    ok = 0
    for i, r in enumerate(results, 1):
        status = "OK" if r.ok else "FAIL"
        if r.ok:
            ok += 1
        hits = ",".join(r.hit_keywords) or "-"
        print(f"[{i}] {status} | {r.name} | {r.phone} | 标红:{hits} | {r.source}")
        print(f"     地址: {r.address[:50]}...")

    out = ROOT / "测试导出结果.xlsx"
    export_to_xlsx(results, out, template_path=ROOT / "5.15新表格.xlsx", hit_color="#FF4444")

    wb = load_workbook(out)
    ws = wb.active
    red_rows = 0
    hit_hex = "FF4444"
    for row_idx in range(2, 2 + len(results)):
        fill = ws[f"A{row_idx}"].fill
        if fill and fill.start_color and fill.start_color.rgb and hit_hex in str(fill.start_color.rgb).upper():
            red_rows += 1

    print(f"\n解析成功: {ok}/{len(results)}")
    print(f"Excel 标红行: {red_rows}")
    print(f"导出文件: {out}")

    if ok < len(records):
        raise SystemExit("存在解析失败项")
    if red_rows < 2:
        raise SystemExit("关键词标红行数不足（期望至少2行：驿站+工商局）")
    print("\n全部测试通过")


if __name__ == "__main__":
    main()
