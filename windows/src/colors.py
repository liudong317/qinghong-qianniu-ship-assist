"""标色配置与颜色工具。"""

from __future__ import annotations

from openpyxl.styles import PatternFill

DEFAULT_HIGHLIGHT_COLORS = {
    "keyword_hit": "#FF4444",
    "parse_warn": "#FFFF99",
}

COLOR_PRESETS = {
    "正红": "#FF0000",
    "浅红": "#FFCCCC",
    "橙色": "#FFA500",
    "黄色": "#FFFF00",
    "绿色": "#90EE90",
    "蓝色": "#ADD8E6",
}


def normalize_hex(color: str, default: str = "#FF4444") -> str:
    if not color:
        return default
    c = color.strip()
    if not c.startswith("#"):
        c = f"#{c}"
    if len(c) == 7:
        return c.upper()
    return default


def hex_to_openpyxl(color: str) -> str:
    return normalize_hex(color).lstrip("#")


def make_fill(color: str, default: str = "#FF4444") -> PatternFill:
    c = hex_to_openpyxl(color or default)
    return PatternFill(start_color=c, end_color=c, fill_type="solid")
