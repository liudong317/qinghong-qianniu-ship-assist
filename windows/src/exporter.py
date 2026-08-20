"""导出千牛标准 xlsx，命中行 A~I 整行标色。"""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .colors import make_fill
from .config import resource_path
from .parser import ParseResult

COLS = "ABCDEFGHI"


def export_to_xlsx(
    results: list[ParseResult],
    output_path: str | Path,
    template_path: str | Path | None = None,
    extract_order_no_to_d: bool = False,
    hit_color: str = "#FF4444",
    warn_color: str = "#FFFF99",
) -> Path:
    template = Path(template_path) if template_path else resource_path("5.15新表格.xlsx")
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    hit_fill = make_fill(hit_color)
    warn_fill = make_fill(warn_color)

    shutil.copy2(template, output)
    wb = load_workbook(output)
    ws = wb.active

    for row_idx in range(2, ws.max_row + 1):
        for col in COLS:
            ws[f"{col}{row_idx}"].value = None
            ws[f"{col}{row_idx}"].fill = PatternFill()

    start_row = 2
    for offset, item in enumerate(results):
        row_idx = start_row + offset
        ws[f"A{row_idx}"] = item.name or ""
        ws[f"B{row_idx}"] = item.phone or ""
        ws[f"C{row_idx}"] = item.address or ""

        order_no = item.order_no if extract_order_no_to_d else ""
        ws[f"D{row_idx}"] = order_no or None
        ws[f"E{row_idx}"] = item.product_info or None
        ws[f"F{row_idx}"] = item.spec_info or None
        ws[f"G{row_idx}"] = item.quantity or None
        ws[f"H{row_idx}"] = item.weight or None
        remark = item.remark or item.error or ""
        ws[f"I{row_idx}"] = remark or None

        fill = None
        if item.hit_keywords:
            fill = hit_fill
        elif item.error or not item.ok:
            fill = warn_fill

        if fill:
            for col in COLS:
                ws[f"{col}{row_idx}"].fill = fill

    wb.save(output)
    return output
