"""批量导入导出、剪贴板、批量替换。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .parser import ParseResult, pick_order_no


def read_text_file(path: str | Path) -> str:
    path = Path(path)
    for enc in ("utf-8-sig", "utf-8", "gbk"):
        try:
            return path.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def import_from_xlsx(path: str | Path) -> list[ParseResult]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    results: list[ParseResult] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        cells = list(row) + [None] * (9 - len(row))
        name = str(cells[0] or "").strip()
        phone = str(cells[1] or "").strip()
        address = str(cells[2] or "").strip()
        order_no = str(cells[3] or "").strip()
        product_info = str(cells[4] or "").strip()
        spec_info = str(cells[5] or "").strip()
        quantity = str(cells[6] or "").strip()
        weight = str(cells[7] or "").strip()
        remark = str(cells[8] or "").strip()
        if not name and not phone and not address:
            continue
        item = ParseResult(
            name=name,
            phone=phone,
            address=address,
            order_no=order_no or pick_order_no(name + address),
            product_info=product_info,
            spec_info=spec_info,
            quantity=quantity,
            weight=weight,
            remark=remark,
            source="import",
            raw=f"{name} {phone} {address}",
        )
        results.append(item)
    return results


def results_to_tsv(
    results: list[ParseResult],
    extract_d: bool = False,
    headers: list[str] | None = None,
) -> str:
    from .template_columns import DEFAULT_TEMPLATE_HEADERS, result_to_row

    hdr = headers or DEFAULT_TEMPLATE_HEADERS
    lines = ["\t".join(hdr)]
    for item in results:
        lines.append("\t".join(result_to_row(item, extract_order_no_to_d=extract_d)))
    return "\n".join(lines)


def batch_replace_in_results(results: list[ParseResult], find: str, replace: str, fields: tuple[str, ...] = ("address",)) -> int:
    if not find:
        return 0
    count = 0
    for item in results:
        for field in fields:
            val = getattr(item, field, "") or ""
            if find in val:
                setattr(item, field, val.replace(find, replace))
                count += 1
    return count
